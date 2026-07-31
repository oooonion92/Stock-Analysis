#!/usr/bin/env python3
"""A 股短线数据采集器：默认采集最近一个已经收盘的交易日。"""

from __future__ import annotations

import argparse
import inspect
import json
import os
import re
import sys
import time
from dataclasses import asdict, dataclass
from datetime import datetime, time as clock_time
from pathlib import Path
from typing import Any, Callable
from zoneinfo import ZoneInfo

import akshare as ak
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


COLLECTOR_VERSION = "3.0.0-latest-completed-trade-date"
DEFAULT_OUTPUT_DIR = Path(r"D:\OneDrive\Stock\短线数据采集")
DATE_PATTERN = re.compile(r"^\d{8}$")
MARKET_TIMEZONE = ZoneInfo("Asia/Shanghai")
DATA_READY_TIME = clock_time(15, 30)

POOL_SCHEMAS = {
    "涨停池": ["序号", "代码", "名称", "涨跌幅", "最新价", "成交额", "流通市值", "总市值", "换手率", "封板资金", "首次封板时间", "最后封板时间", "炸板次数", "涨停统计", "连板数", "所属行业"],
    "炸板池": ["序号", "代码", "名称", "涨跌幅", "最新价", "涨停价", "成交额", "流通市值", "总市值", "换手率", "涨速", "首次封板时间", "炸板次数", "涨停统计", "振幅", "所属行业"],
    "跌停池": ["序号", "代码", "名称", "涨跌幅", "最新价", "成交额", "流通市值", "总市值", "动态市盈率", "换手率", "封单资金", "最后封板时间", "板上成交额", "连续跌停", "开板次数", "所属行业"],
    "昨日涨停反馈": ["序号", "代码", "名称", "涨跌幅", "最新价", "涨停价", "成交额", "流通市值", "总市值", "换手率", "涨速", "振幅", "昨日封板时间", "昨日连板数", "涨停统计", "所属行业"],
}


@dataclass
class SourceStatus:
    source: str
    ok: bool
    state: str
    rows: int | None
    columns: list[str] | None = None
    error: str | None = None


def validate_date(value: str) -> str:
    if not DATE_PATTERN.fullmatch(value):
        raise ValueError(f"日期必须为 YYYYMMDD：{value!r}")
    datetime.strptime(value, "%Y%m%d")
    return value


def fetch_with_retry(
    label: str,
    source: str,
    fetcher: Callable[[], pd.DataFrame],
    attempts: int = 3,
    pause_seconds: float = 1.5,
) -> tuple[pd.DataFrame | None, SourceStatus]:
    """成功空集和请求失败严格分开；失败绝不转换为 0。"""
    last_error: Exception | None = None
    for attempt in range(1, attempts + 1):
        try:
            frame = fetcher()
            if frame is None:
                frame = pd.DataFrame()
            if not isinstance(frame, pd.DataFrame):
                raise TypeError(f"返回 {type(frame).__name__}，预期 DataFrame")
            state = "success_empty" if frame.empty else "success"
            return frame, SourceStatus(source, True, state, len(frame), [str(x) for x in frame.columns])
        except Exception as exc:
            last_error = exc
            if attempt < attempts:
                time.sleep(pause_seconds * attempt)
    message = f"{label}获取失败：{type(last_error).__name__}: {last_error}"
    return None, SourceStatus(source, False, "failed", None, None, message)


def find_column(frame: pd.DataFrame | None, *candidates: str) -> str | None:
    if frame is not None:
        for name in candidates:
            if name in frame.columns:
                return name
    return None


def normalize_code(value: Any) -> str | pd.NA:
    if value is None or pd.isna(value):
        return pd.NA
    match = re.search(r"(\d{1,6})", str(value).strip())
    return match.group(1).zfill(6) if match else pd.NA


def normalize_time(value: Any) -> str | pd.NA:
    if value is None or pd.isna(value):
        return pd.NA
    digits = re.sub(r"\D", "", str(value))
    return digits.zfill(6)[-6:] if digits else pd.NA


def normalize_frame(frame: pd.DataFrame | None, schema_name: str) -> pd.DataFrame | None:
    if frame is None:
        return None
    work = frame.copy()
    if work.empty and len(work.columns) == 0:
        work = pd.DataFrame(columns=POOL_SCHEMAS[schema_name])
    code_col = find_column(work, "代码", "证券代码", "股票代码")
    if code_col:
        work[code_col] = work[code_col].map(normalize_code).astype("string")
    for col in ("首次封板时间", "最后封板时间", "昨日封板时间"):
        if col in work.columns:
            work[col] = work[col].map(normalize_time).astype("string")
    return work


def code_series(frame: pd.DataFrame | None) -> pd.Series:
    if frame is None:
        return pd.Series(dtype="string")
    column = find_column(frame, "代码", "证券代码", "股票代码")
    if column is None:
        return pd.Series(pd.NA, index=frame.index, dtype="string")
    return frame[column].map(normalize_code).astype("string")


def numeric_series(frame: pd.DataFrame | None, *candidates: str) -> pd.Series:
    column = find_column(frame, *candidates)
    if frame is None or column is None:
        return pd.Series(dtype="float64")
    return pd.to_numeric(frame[column], errors="coerce")


def safe_round(value: Any, digits: int = 2) -> float | None:
    if value is None or pd.isna(value):
        return None
    return round(float(value), digits)


def ratio(numerator: int | float | None, denominator: int | float | None) -> float | None:
    if numerator is None or denominator in (None, 0):
        return None
    return round(float(numerator) / float(denominator) * 100, 2)


def board_count(value: Any, prefer_denominator: bool = False) -> int | None:
    """解析连板数；对“5/3、5天3板”取板数 3，不误取天数 5。"""
    if value is None or pd.isna(value):
        return None
    text = str(value).strip()
    if re.fullmatch(r"\d+(?:\.0+)?", text):
        return int(float(text))
    slash = re.search(r"(\d+)\s*/\s*(\d+)", text)
    if slash:
        return int(slash.group(2) if prefer_denominator else slash.group(2))
    board = re.search(r"(\d+)\s*板", text)
    return int(board.group(1)) if board else None


def tier_series(frame: pd.DataFrame | None, previous: bool = False) -> pd.Series:
    candidates = ("昨日连板数", "连板数") if previous else ("连板数",)
    column = find_column(frame, *candidates)
    if frame is None or column is None:
        return pd.Series(pd.NA, index=frame.index if frame is not None else None, dtype="Int64")
    return frame[column].map(board_count).astype("Int64")


def previous_trade_date(trade_date: str) -> tuple[str | None, SourceStatus]:
    calendar, status = fetch_with_retry(
        "交易日历", "新浪/tool_trade_date_hist_sina", ak.tool_trade_date_hist_sina
    )
    if calendar is None:
        return None, status
    if "trade_date" not in calendar.columns:
        return None, SourceStatus(status.source, False, "failed", len(calendar), list(calendar.columns), "交易日历缺少 trade_date 字段")
    target = pd.Timestamp(datetime.strptime(trade_date, "%Y%m%d").date())
    dates = pd.to_datetime(calendar["trade_date"], errors="coerce").dropna()
    previous = dates[dates < target]
    if previous.empty:
        return None, SourceStatus(status.source, False, "failed", len(calendar), list(calendar.columns), f"找不到 {trade_date} 的上一交易日")
    return previous.max().strftime("%Y%m%d"), status


def latest_completed_trade_date(now: datetime | None = None) -> tuple[str, SourceStatus, str]:
    """选择最近已收盘交易日；交易日 15:30 前仍使用上一交易日。"""
    calendar, status = fetch_with_retry(
        "交易日历", "新浪/tool_trade_date_hist_sina", ak.tool_trade_date_hist_sina
    )
    if calendar is None:
        raise RuntimeError(status.error or "交易日历获取失败")
    if "trade_date" not in calendar.columns:
        raise RuntimeError("交易日历缺少 trade_date 字段")
    current = now.astimezone(MARKET_TIMEZONE) if now else datetime.now(MARKET_TIMEZONE)
    dates = pd.to_datetime(calendar["trade_date"], errors="coerce").dropna().dt.normalize()
    today = pd.Timestamp(current.date())
    today_is_trade_date = bool(dates.eq(today).any())
    include_today = today_is_trade_date and current.time().replace(tzinfo=None) >= DATA_READY_TIME
    eligible = dates[dates.le(today) if include_today else dates.lt(today)]
    if eligible.empty:
        raise RuntimeError(f"交易日历中找不到 {today:%Y-%m-%d} 之前已完成的交易日")
    selected = eligible.max().strftime("%Y%m%d")
    if include_today:
        reason = f"当前为交易日且已过数据就绪时间 {DATA_READY_TIME.strftime('%H:%M')}"
    elif today_is_trade_date:
        reason = f"当前交易日尚未到数据就绪时间 {DATA_READY_TIME.strftime('%H:%M')}，回退上一交易日"
    else:
        reason = "当前为周末或非交易日，回退最近交易日"
    return selected, status, reason


def promotion_rates(previous_zt: pd.DataFrame | None, current_zt: pd.DataFrame | None) -> list[dict[str, Any]]:
    labels = (("1进2", 1, 1), ("2进3", 2, 2), ("3板及以上", 3, None))
    if previous_zt is None or current_zt is None:
        return [{"层级": x[0], "晋级数": None, "昨日样本数": None, "晋级率_pct": None, "状态": "数据源失败"} for x in labels]
    previous = previous_zt.copy()
    previous["_代码"] = code_series(previous)
    previous["_昨日层级"] = tier_series(previous)
    current = current_zt.copy()
    current["_代码"] = code_series(current)
    current["_今日层级"] = tier_series(current)
    current_tiers = dict(zip(current["_代码"].dropna(), current.loc[current["_代码"].notna(), "_今日层级"]))
    result = []
    for label, lower, upper in labels:
        cohort = previous[previous["_昨日层级"].ge(lower).fillna(False)]
        if upper is not None:
            cohort = cohort[cohort["_昨日层级"].le(upper).fillna(False)]
        denominator = int(len(cohort))
        numerator = 0
        promoted_codes = []
        for _, row in cohort.iterrows():
            code = row["_代码"]
            current_tier = current_tiers.get(code)
            if pd.notna(current_tier) and int(current_tier) == int(row["_昨日层级"]) + 1:
                numerator += 1
                promoted_codes.append(code)
        result.append({"层级": label, "晋级数": numerator, "昨日样本数": denominator, "晋级率_pct": ratio(numerator, denominator), "晋级代码": promoted_codes, "状态": "成功"})
    return result


def ladder(zt_pool: pd.DataFrame | None) -> dict[str, Any]:
    if zt_pool is None:
        return {"最高连板": None, "连板家数": None, "梯队": [], "状态": "数据源失败"}
    work = zt_pool.copy()
    work["_层级"] = tier_series(work)
    work["_代码"] = code_series(work)
    name_col = find_column(work, "名称", "股票简称")
    work = work[work["_层级"].ge(2).fillna(False)]
    tiers = []
    for level in sorted(work["_层级"].dropna().astype(int).unique(), reverse=True):
        rows = work[work["_层级"].eq(level)]
        stocks = [{"代码": row["_代码"], "名称": row.get(name_col) if name_col else None} for _, row in rows.iterrows()]
        tiers.append({"连板数": int(level), "家数": len(stocks), "股票": stocks})
    return {"最高连板": int(work["_层级"].max()) if not work.empty else 0, "连板家数": int(len(work)), "梯队": tiers, "状态": "成功"}


def first_boards(zt_pool: pd.DataFrame | None) -> pd.DataFrame | None:
    if zt_pool is None:
        return None
    return zt_pool.loc[tier_series(zt_pool).eq(1).fillna(False)].copy()


def time_values(frame: pd.DataFrame, column: str) -> pd.Series:
    if column not in frame.columns:
        return pd.Series(dtype="string")
    values = frame[column].map(normalize_time).astype("string")
    return values[values.str.fullmatch(r"\d{6}", na=False)]


def seal_quality(zt_pool: pd.DataFrame | None) -> dict[str, Any]:
    if zt_pool is None:
        return {"样本数": None, "状态": "数据源失败"}
    broken = numeric_series(zt_pool, "炸板次数").dropna()
    funds = numeric_series(zt_pool, "封板资金").dropna()
    first = time_values(zt_pool, "首次封板时间")
    last = time_values(zt_pool, "最后封板时间")
    never_broken = int(broken.eq(0).sum()) if "炸板次数" in zt_pool.columns else None
    return {
        "样本数": int(len(zt_pool)),
        "状态": "成功空集" if zt_pool.empty else "成功",
        "未炸板家数": never_broken,
        "未炸板率_pct": ratio(never_broken, len(zt_pool)) if never_broken is not None else None,
        "炸板次数合计": int(broken.sum()) if not broken.empty else (0 if "炸板次数" in zt_pool.columns else None),
        "平均炸板次数": safe_round(broken.mean()),
        "封板资金合计": safe_round(funds.sum()),
        "封板资金中位数": safe_round(funds.median()),
        "最早首次封板时间": first.min() if not first.empty else None,
        "首次封板时间中位数": first.sort_values().iloc[len(first) // 2] if not first.empty else None,
        "最晚最后封板时间": last.max() if not last.empty else None,
        "最后封板时间中位数": last.sort_values().iloc[len(last) // 2] if not last.empty else None,
        "10点前首封家数": int(first.lt("100000").sum()) if not first.empty else (0 if "首次封板时间" in zt_pool.columns else None),
        "14点后首封家数": int(first.ge("140000").sum()) if not first.empty else (0 if "首次封板时间" in zt_pool.columns else None),
    }


def feedback_metrics(feedback: pd.DataFrame | None, current_zt: pd.DataFrame | None) -> dict[str, Any]:
    if feedback is None:
        return {"样本数": None, "状态": "数据源失败"}
    change = numeric_series(feedback, "涨跌幅").dropna()
    current_codes = set(code_series(current_zt).dropna()) if current_zt is not None else None
    again_count = int(code_series(feedback).isin(current_codes).sum()) if current_codes is not None else None
    return {
        "样本数": int(len(feedback)),
        "状态": "成功空集" if feedback.empty else "成功",
        "有效涨跌幅样本数": int(len(change)),
        "平均涨跌幅_pct": safe_round(change.mean()),
        "中位数涨跌幅_pct": safe_round(change.median()),
        "上涨率_pct": ratio(int(change.gt(0).sum()), len(change)) if len(change) else None,
        "再次涨停家数": again_count,
        "再次涨停率_pct": ratio(again_count, len(feedback)),
        "跌超5pct家数": int(change.le(-5).sum()) if len(change) else 0,
        "跌超7pct家数": int(change.le(-7).sum()) if len(change) else 0,
        "最差涨跌幅_pct": safe_round(change.min()),
    }


def complete_feedback_cohort(
    previous_zt: pd.DataFrame | None, feedback: pd.DataFrame | None
) -> tuple[pd.DataFrame | None, list[str]]:
    """用上一交易日涨停池补齐反馈接口漏掉的股票；缺失行情保持空值。"""
    if previous_zt is None or feedback is None:
        return feedback, []
    work = feedback.copy()
    work["反馈数据状态"] = "接口返回"
    feedback_codes = set(code_series(work).dropna())
    previous = previous_zt.copy()
    previous["_代码"] = code_series(previous)
    missing = previous[~previous["_代码"].isin(feedback_codes)]
    if missing.empty:
        return work, []
    rows = []
    for _, row in missing.iterrows():
        record = {column: pd.NA for column in work.columns}
        record["代码"] = row["_代码"]
        if "名称" in work.columns:
            record["名称"] = row.get("名称")
        if "昨日连板数" in work.columns:
            record["昨日连板数"] = row.get("连板数")
        if "昨日封板时间" in work.columns:
            record["昨日封板时间"] = row.get("最后封板时间")
        if "涨停统计" in work.columns:
            record["涨停统计"] = row.get("涨停统计")
        if "所属行业" in work.columns:
            record["所属行业"] = row.get("所属行业")
        record["反馈数据状态"] = "反馈接口缺失"
        rows.append(record)
    completed = pd.concat([work, pd.DataFrame(rows)], ignore_index=True)
    completed["代码"] = completed["代码"].astype("string").str.zfill(6)
    return completed, sorted(missing["_代码"].dropna().tolist())


def json_default(value: Any) -> Any:
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.isoformat()
    if value is pd.NA or (not isinstance(value, (list, dict)) and pd.isna(value)):
        return None
    if hasattr(value, "item"):
        return value.item()
    return str(value)


def atomic_csv(frame: pd.DataFrame, path: Path) -> None:
    temp = path.with_suffix(path.suffix + ".tmp")
    frame.to_csv(temp, index=False, encoding="utf-8-sig")
    os.replace(temp, path)


def atomic_json(payload: dict[str, Any], path: Path) -> None:
    temp = path.with_suffix(path.suffix + ".tmp")
    temp.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=json_default, allow_nan=False), encoding="utf-8")
    os.replace(temp, path)


def export_frame(frame: pd.DataFrame | None, schema_name: str, trade_date: str) -> pd.DataFrame:
    work = frame.copy() if frame is not None else pd.DataFrame(columns=POOL_SCHEMAS[schema_name])
    if "交易日期" not in work.columns:
        work.insert(0, "交易日期", trade_date)
    work["交易日期"] = work["交易日期"].astype("string")
    code_col = find_column(work, "代码", "证券代码", "股票代码")
    if code_col:
        work[code_col] = work[code_col].astype("string").str.zfill(6)
    return work


def atomic_excel(frames: dict[str, pd.DataFrame], snapshot: dict[str, Any], path: Path) -> None:
    temp = path.with_suffix(".tmp.xlsx")
    overview = pd.DataFrame([snapshot["市场概览"]])
    statuses = pd.DataFrame([{"数据集": name, **status} for name, status in snapshot["数据源状态"].items()])
    promotions = pd.DataFrame(snapshot["晋级率"])
    if "晋级代码" in promotions.columns:
        promotions["晋级代码"] = promotions["晋级代码"].map(
            lambda x: "代码 " + ",".join(x) if isinstance(x, list) and x else ""
        )
    ladder_rows = []
    for tier in snapshot["连板梯队"]["梯队"]:
        ladder_rows.extend({"连板数": tier["连板数"], **stock} for stock in tier["股票"])
    with pd.ExcelWriter(temp, engine="openpyxl") as writer:
        overview.to_excel(writer, sheet_name="概览", index=False)
        statuses.to_excel(writer, sheet_name="数据源状态", index=False)
        promotions.to_excel(writer, sheet_name="晋级率", index=False)
        pd.DataFrame(ladder_rows, columns=["连板数", "代码", "名称"]).to_excel(writer, sheet_name="连板梯队", index=False)
        for sheet, frame in frames.items():
            frame.to_excel(writer, sheet_name=sheet, index=False)
    workbook = load_workbook(temp)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    success_fill = PatternFill("solid", fgColor="E2F0D9")
    warning_fill = PatternFill("solid", fgColor="FFF2CC")
    failed_fill = PatternFill("solid", fgColor="FCE4D6")
    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.row_dimensions[1].height = 24
        headers = {cell.value: cell.column for cell in sheet[1] if cell.value is not None}
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column_cells in sheet.iter_cols(1, sheet.max_column):
            letter = get_column_letter(column_cells[0].column)
            header = str(column_cells[0].value or "")
            max_length = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=8)
            sheet.column_dimensions[letter].width = min(max(max_length + 2, 10), 28)
            for cell in column_cells[1:]:
                cell.alignment = Alignment(vertical="center", wrap_text=False)
                if header == "交易日期":
                    cell.number_format = "0"
                elif header == "代码":
                    cell.number_format = "000000"
                elif "时间" in header:
                    normalized = normalize_time(cell.value)
                    if normalized is not pd.NA and not pd.isna(normalized):
                        cell.value = str(normalized)
                    cell.number_format = "@"
                    cell.quotePrefix = True
                elif "_pct" in header or "率" in header or "涨跌幅" in header:
                    cell.number_format = "0.00"
                elif any(word in header for word in ("资金", "成交额", "市值", "最新价", "涨停价")):
                    cell.number_format = "#,##0.00"
        state_column = headers.get("state") or headers.get("状态")
        if state_column:
            for row in range(2, sheet.max_row + 1):
                cell = sheet.cell(row, state_column)
                text = str(cell.value or "").lower()
                cell.fill = failed_fill if "fail" in text or "失败" in text else warning_fill if "empty" in text or "空" in text or "缺失" in text else success_fill
        for long_header in ("columns", "error"):
            long_column = headers.get(long_header)
            if long_column:
                sheet.column_dimensions[get_column_letter(long_column)].width = 28
                for row in range(2, sheet.max_row + 1):
                    sheet.cell(row, long_column).alignment = Alignment(vertical="top", wrap_text=True)
                    sheet.row_dimensions[row].height = 42
        if sheet.title == "数据源状态":
            for header, width in (("数据集", 22), ("source", 42), ("columns", 45), ("error", 36)):
                column = headers.get(header)
                if column:
                    sheet.column_dimensions[get_column_letter(column)].width = width
            for row in range(2, sheet.max_row + 1):
                sheet.row_dimensions[row].height = 60
        if sheet.title == "晋级率" and headers.get("晋级代码"):
            sheet.column_dimensions[get_column_letter(headers["晋级代码"])].width = 55
        sheet.auto_filter.ref = sheet.dimensions
    workbook.save(temp)
    os.replace(temp, path)


def upsert_summary(row: dict[str, Any], path: Path) -> None:
    new_row = pd.DataFrame([row])
    if path.exists():
        old = pd.read_csv(path, dtype={"交易日期": "string"})
        combined = pd.concat([old, new_row], ignore_index=True)
    else:
        combined = new_row
    combined["交易日期"] = combined["交易日期"].astype("string")
    combined = combined.drop_duplicates("交易日期", keep="last").sort_values("交易日期")
    atomic_csv(combined, path)


def promotion_map(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    return {row["层级"]: row for row in rows}


def collect_one_day(trade_date: str, output_root: Path) -> dict[str, Any]:
    trade_date = validate_date(trade_date)
    print(f"\n[{trade_date}] 开始采集")
    specs = {
        "涨停池": ("东方财富/stock_zt_pool_em", lambda: ak.stock_zt_pool_em(date=trade_date)),
        "炸板池": ("东方财富/stock_zt_pool_zbgc_em", lambda: ak.stock_zt_pool_zbgc_em(date=trade_date)),
        "跌停池": ("东方财富/stock_zt_pool_dtgc_em", lambda: ak.stock_zt_pool_dtgc_em(date=trade_date)),
        "昨日涨停反馈": ("东方财富/stock_zt_pool_previous_em", lambda: ak.stock_zt_pool_previous_em(date=trade_date)),
    }
    frames: dict[str, pd.DataFrame | None] = {}
    statuses: dict[str, SourceStatus] = {}
    for label, (source, fetcher) in specs.items():
        raw, status = fetch_with_retry(label, source, fetcher)
        frames[label] = normalize_frame(raw, label)
        statuses[label] = status
        print(f"[{trade_date}] {label}: {status.state}, rows={status.rows}")

    previous_date, calendar_status = previous_trade_date(trade_date)
    statuses["交易日历"] = calendar_status
    previous_zt = None
    if previous_date:
        previous_zt, previous_status = fetch_with_retry(
            "上一交易日涨停池", "东方财富/stock_zt_pool_em", lambda: ak.stock_zt_pool_em(date=previous_date)
        )
        previous_zt = normalize_frame(previous_zt, "涨停池")
    else:
        previous_status = SourceStatus("东方财富/stock_zt_pool_em", False, "failed", None, None, "上一交易日不可用")
    statuses["上一交易日涨停池"] = previous_status

    zt, zb, dt, raw_feedback = frames["涨停池"], frames["炸板池"], frames["跌停池"], frames["昨日涨停反馈"]
    zt_count = len(zt) if zt is not None else None
    zb_count = len(zb) if zb is not None else None
    dt_count = len(dt) if dt is not None else None
    denominator = zt_count + zb_count if zt_count is not None and zb_count is not None else None
    promotions = promotion_rates(previous_zt, zt)
    ladder_data = ladder(zt)
    first = first_boards(zt)
    first_count = len(first) if first is not None else None
    feedback, feedback_missing_codes = complete_feedback_cohort(previous_zt, raw_feedback)
    frames["昨日涨停反馈"] = feedback
    previous_codes_match = None
    if previous_zt is not None and raw_feedback is not None:
        previous_codes_match = set(code_series(previous_zt).dropna()) == set(code_series(raw_feedback).dropna())

    snapshot = {
        "schema_version": "2.0",
        "collector_version": COLLECTOR_VERSION,
        "python_version": sys.version.split()[0],
        "akshare_version": getattr(ak, "__version__", None),
        "pandas_version": pd.__version__,
        "交易日期": trade_date,
        "上一交易日": previous_date,
        "采集时间": datetime.now().astimezone().isoformat(timespec="seconds"),
        "数据源状态": {name: asdict(status) for name, status in statuses.items()},
        "接口审计": {
            "历史日期参数": {name: "date" in inspect.signature(getattr(ak, source.split("/")[-1])).parameters for name, source in [("涨停池", specs["涨停池"][0]), ("炸板池", specs["炸板池"][0]), ("跌停池", specs["跌停池"][0]), ("昨日涨停反馈", specs["昨日涨停反馈"][0])]},
            "昨日反馈与上一交易日涨停代码集合一致": previous_codes_match,
            "昨日反馈接口缺失代码": feedback_missing_codes,
            "昨日反馈输出已按上一交易日涨停池补齐": bool(feedback_missing_codes),
        },
        "市场概览": {
            "涨停家数": zt_count,
            "炸板家数": zb_count,
            "跌停家数": dt_count,
            "封板率_pct": ratio(zt_count, denominator),
            "炸板率_pct": ratio(zb_count, denominator),
            "首板家数": first_count,
            "最高连板": ladder_data["最高连板"],
            "连板家数": ladder_data["连板家数"],
        },
        "封板质量": seal_quality(zt),
        "晋级率": promotions,
        "连板梯队": ladder_data,
        "昨日涨停反馈": feedback_metrics(feedback, zt),
        "口径说明": [
            "炸板率=炸板家数/(最终涨停家数+炸板家数)，炸板家数取炸板池明细行数。",
            "晋级率分母为上一交易日对应连板层级，分子要求当日继续涨停且连板数恰好加一。",
            "接口失败时指标为 null；success_empty 表示接口成功但结果为空。",
            "未调用实时市场活跃度或实时指数接口，也未回填全市场涨跌家数。",
        ],
    }

    day_dir = output_root / trade_date
    day_dir.mkdir(parents=True, exist_ok=True)
    export_frames = {
        "涨停池": export_frame(zt, "涨停池", trade_date),
        "炸板池": export_frame(zb, "炸板池", trade_date),
        "跌停池": export_frame(dt, "跌停池", trade_date),
        "昨日涨停反馈": export_frame(feedback, "昨日涨停反馈", trade_date),
        "首板": export_frame(first, "涨停池", trade_date),
    }
    filenames = {"涨停池": "zt_pool.csv", "炸板池": "zb_pool.csv", "跌停池": "dt_pool.csv", "昨日涨停反馈": "previous_zt_feedback.csv", "首板": "first_boards.csv"}
    for label, frame in export_frames.items():
        atomic_csv(frame, day_dir / filenames[label])
    atomic_json(snapshot, day_dir / "snapshot.json")
    atomic_excel(export_frames, snapshot, day_dir / f"短线数据_{trade_date}.xlsx")

    pm = promotion_map(promotions)
    fb = snapshot["昨日涨停反馈"]
    summary = {
        "交易日期": trade_date,
        **snapshot["市场概览"],
        "1进2_晋级数": pm["1进2"]["晋级数"], "1进2_昨日样本数": pm["1进2"]["昨日样本数"], "1进2_晋级率_pct": pm["1进2"]["晋级率_pct"],
        "2进3_晋级数": pm["2进3"]["晋级数"], "2进3_昨日样本数": pm["2进3"]["昨日样本数"], "2进3_晋级率_pct": pm["2进3"]["晋级率_pct"],
        "3板及以上_晋级数": pm["3板及以上"]["晋级数"], "3板及以上_昨日样本数": pm["3板及以上"]["昨日样本数"], "3板及以上_晋级率_pct": pm["3板及以上"]["晋级率_pct"],
        "昨日涨停平均涨跌幅_pct": fb.get("平均涨跌幅_pct"),
        "昨日涨停上涨率_pct": fb.get("上涨率_pct"),
        "昨日涨停再次涨停率_pct": fb.get("再次涨停率_pct"),
        "涨停池状态": statuses["涨停池"].state, "炸板池状态": statuses["炸板池"].state, "跌停池状态": statuses["跌停池"].state,
        "采集是否完整": all(x.ok for x in statuses.values()),
    }
    upsert_summary(summary, output_root / "DB_Short_Term_Summary.csv")
    print(f"[{trade_date}] 已保存：{day_dir}")
    return snapshot


def read_csv(path: Path) -> pd.DataFrame:
    return pd.read_csv(path, dtype={"交易日期": "string", "代码": "string", "证券代码": "string", "股票代码": "string"})


def validate_outputs(output_root: Path, trade_dates: list[str]) -> list[dict[str, Any]]:
    results = []
    for trade_date in trade_dates:
        day_dir = output_root / trade_date
        required = ["zt_pool.csv", "zb_pool.csv", "dt_pool.csv", "first_boards.csv", "previous_zt_feedback.csv", "snapshot.json", f"短线数据_{trade_date}.xlsx"]
        missing = [name for name in required if not (day_dir / name).exists()]
        if missing:
            raise AssertionError(f"{trade_date} 缺少文件：{missing}")
        zt, zb, dt, first, feedback = [read_csv(day_dir / name) for name in required[:5]]
        snapshot = json.loads((day_dir / "snapshot.json").read_text(encoding="utf-8"))
        sheets = pd.read_excel(day_dir / f"短线数据_{trade_date}.xlsx", sheet_name=None, dtype={"代码": "string"})
        workbook = load_workbook(day_dir / f"短线数据_{trade_date}.xlsx", read_only=False, data_only=False)
        for worksheet in workbook.worksheets:
            headers = {cell.value: cell.column for cell in worksheet[1] if cell.value is not None}
            for header in ("首次封板时间", "最后封板时间", "昨日封板时间"):
                column = headers.get(header)
                if not column:
                    continue
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row, column)
                    if cell.value is None:
                        continue
                    if not (
                        cell.data_type == "s"
                        and re.fullmatch(r"\d{6}", str(cell.value))
                        and cell.number_format == "@"
                        and cell.quotePrefix
                    ):
                        raise AssertionError(
                            f"{trade_date} {worksheet.title}/{header} 不是六位 Excel 文本："
                            f"value={cell.value!r}, type={cell.data_type}, format={cell.number_format!r}"
                        )
        for label, frame in (("涨停", zt), ("炸板", zb), ("跌停", dt), ("首板", first), ("昨日反馈", feedback)):
            if len(frame) and (frame["交易日期"].isna().any() or not frame["交易日期"].eq(trade_date).all()):
                raise AssertionError(f"{trade_date} {label}存在空日期或错误日期")
            if "代码" in frame.columns and not frame["代码"].dropna().str.fullmatch(r"\d{6}").all():
                raise AssertionError(f"{trade_date} {label}存在非六位代码")
        overview = snapshot["市场概览"]
        if (len(zt), len(zb), len(dt), len(first)) != (overview["涨停家数"], overview["炸板家数"], overview["跌停家数"], overview["首板家数"]):
            raise AssertionError(f"{trade_date} 明细行数与 snapshot 不一致")
        if len(first) != int(pd.to_numeric(zt.get("连板数", pd.Series(dtype=float)), errors="coerce").eq(1).sum()):
            raise AssertionError(f"{trade_date} 首板数量不可由涨停池复算")
        rate_sum = None if overview["封板率_pct"] is None or overview["炸板率_pct"] is None else round(overview["封板率_pct"] + overview["炸板率_pct"], 2)
        if rate_sum is not None and abs(rate_sum - 100) > 0.01:
            raise AssertionError(f"{trade_date} 封板率与炸板率合计不是 100%")
        previous_codes = set(feedback["代码"].dropna()) if "代码" in feedback else set()
        for row in snapshot["晋级率"]:
            if row["昨日样本数"] is not None and row["晋级率_pct"] != ratio(row["晋级数"], row["昨日样本数"]):
                raise AssertionError(f"{trade_date} {row['层级']} 晋级率不可复算")
            if any(code not in set(zt["代码"].dropna()) or code not in previous_codes for code in row.get("晋级代码", [])):
                raise AssertionError(f"{trade_date} {row['层级']} 晋级代码不在两日交集")
        results.append({"交易日期": trade_date, "文件完整": True, "Excel工作表数": len(sheets), "涨停": len(zt), "炸板": len(zb), "跌停": len(dt), "首板": len(first), "封板率与炸板率合计": rate_sum, "六位代码": True, "晋级率可复算": True})
    summary = read_csv(output_root / "DB_Short_Term_Summary.csv")
    if summary["交易日期"].duplicated().any():
        raise AssertionError("汇总表存在重复交易日期")
    missing_dates = set(trade_dates) - set(summary["交易日期"])
    if missing_dates:
        raise AssertionError(f"汇总表缺少本次交易日期：{sorted(missing_dates)}")
    return results


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="最近已收盘交易日 A 股短线数据采集器")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--trade-date", type=validate_date, help="手工指定 YYYYMMDD；默认自动选择最近已收盘交易日")
    parser.add_argument("--verify-only", action="store_true", help="只校验已生成文件")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    args.output_dir.mkdir(parents=True, exist_ok=True)
    if args.trade_date:
        trade_date = args.trade_date
        selection_reason = "使用命令行手工指定日期"
    else:
        trade_date, _, selection_reason = latest_completed_trade_date()
    trade_dates = [trade_date]
    print(f"本次采集日期：{trade_date}")
    print(f"日期选择说明：{selection_reason}")
    print(f"输出目录：{args.output_dir}")
    if not args.verify_only:
        collect_one_day(trade_date, args.output_dir)
    checks = validate_outputs(args.output_dir, trade_dates)
    print("\n验证结果：")
    for row in checks:
        print(json.dumps(row, ensure_ascii=False))
    print(f"交易日 {trade_date} 采集与校验完成。")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
